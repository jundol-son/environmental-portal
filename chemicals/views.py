import csv
from datetime import date
from zipfile import BadZipFile

from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .forms import CompletionCodeForm, TrainingUploadForm
from .models import CompletionSubmissionLog, HandlerProfile, NicsNotice, TrainingCompletion
from .utils import crawl_nics_notices
from django.contrib import messages

@login_required
def chemical_check(request):
    # 연동할 외부 URL (예: 화학물질안전원 또는 자체 법령 시스템)
    external_url = "https://www.safetydata.go.kr/" # 실제 필요한 URL로 교체하세요
    return render(request, 'chemicals/external_viewer.html', {'external_url': external_url})

def nics_notice_list(request):
    # DB에 저장된 고시 목록을 가져옴 (최신순)
    notices = NicsNotice.objects.all()
    return render(request, 'chemicals/nics_list.html', {'notices': notices})

def nics_notice_list(request):
    # 'update' 파라미터가 들어오면 크롤링 실행
    if 'update' in request.GET:
        count = crawl_nics_notices()
        messages.success(request, f"{count}건의 새로운 고시가 업데이트되었습니다.")
        return redirect('nics_notice_list')

    notices = NicsNotice.objects.all().order_by('-reg_date')
    return render(request, 'chemicals/nics_list.html', {'notices': notices})


def _mock_knoxid(user):
    # ponytail: local SSO mock; replace this return value when the SSO field is confirmed.
    return user.username


def _year_from_request(request):
    try:
        year = int(request.GET.get("year", date.today().year))
    except (TypeError, ValueError):
        return date.today().year
    return year if 2000 <= year <= 2100 else date.today().year


def _filtered_completions(request, target_year):
    completions = TrainingCompletion.objects.select_related("handler").filter(
        target_year=target_year,
        handler__is_active=True,
    )
    department = request.GET.get("department", "").strip()
    if department:
        completions = completions.filter(handler__department=department)
    name = request.GET.get("name", "").strip()
    if name:
        completions = completions.filter(handler__name__icontains=name)
    status = request.GET.get("status")
    if status == "completed":
        completions = completions.filter(is_completed=True)
    elif status == "pending":
        completions = completions.filter(is_completed=False)
    return completions


@login_required
def training_dashboard(request):
    target_year = _year_from_request(request)
    can_manage = request.user.has_perm("chemicals.change_trainingcompletion")
    viewer_knoxid = _mock_knoxid(request.user)
    viewer_profile = HandlerProfile.objects.filter(
        knoxid=viewer_knoxid,
        is_active=True,
    ).first()

    base_completions = TrainingCompletion.objects.select_related("handler").filter(
        target_year=target_year,
        handler__is_active=True,
    )
    if not can_manage:
        base_completions = base_completions.filter(handler__knoxid=viewer_knoxid)

    selected_department = request.GET.get("department", "").strip()
    if can_manage and selected_department:
        base_completions = base_completions.filter(handler__department=selected_department)
    selected_name = request.GET.get("name", "").strip()
    if can_manage and selected_name:
        base_completions = base_completions.filter(handler__name__icontains=selected_name)

    total_count = base_completions.count()
    completed_count = base_completions.filter(is_completed=True).count()
    pending_count = total_count - completed_count
    completion_rate = round(completed_count * 100 / total_count, 1) if total_count else 0

    completions = base_completions
    selected_status = request.GET.get("status", "")
    if selected_status == "completed":
        completions = completions.filter(is_completed=True)
    elif selected_status == "pending":
        completions = completions.filter(is_completed=False)

    all_years = set(
        TrainingCompletion.objects.values_list("target_year", flat=True).distinct()
        if can_manage
        else TrainingCompletion.objects.filter(
            handler__knoxid=viewer_knoxid
        ).values_list("target_year", flat=True).distinct()
    )
    all_years.add(date.today().year)

    departments = []
    department_stats = []
    if can_manage:
        departments = list(
            TrainingCompletion.objects.filter(
                target_year=target_year,
                handler__is_active=True,
            )
            .exclude(handler__department="")
            .values_list("handler__department", flat=True)
            .distinct()
            .order_by("handler__department")
        )
        department_stats = list(
            TrainingCompletion.objects.filter(
                target_year=target_year,
                handler__is_active=True,
            )
            .values("handler__department")
            .annotate(
                total=Count("id"),
                completed=Count("id", filter=Q(is_completed=True)),
            )
            .order_by("handler__department")
        )
        for stat in department_stats:
            stat["rate"] = round(stat["completed"] * 100 / stat["total"], 1)

    context = {
        "can_manage": can_manage,
        "target_year": target_year,
        "years": sorted(all_years, reverse=True),
        "departments": departments,
        "department_stats": department_stats,
        "selected_department": selected_department,
        "selected_name": selected_name,
        "selected_status": selected_status,
        "total_count": total_count,
        "completed_count": completed_count,
        "pending_count": pending_count,
        "completion_rate": completion_rate,
        "completions": Paginator(completions, 50).get_page(request.GET.get("page")),
        "viewer_knoxid": viewer_knoxid,
        "viewer_profile": viewer_profile,
    }
    return render(request, "chemicals/training_dashboard.html", context)


@login_required
@require_POST
def training_submit(request, completion_id):
    completions = TrainingCompletion.objects.select_related("handler").filter(
        handler__is_active=True
    )
    if not request.user.has_perm("chemicals.change_trainingcompletion"):
        completions = completions.filter(handler__knoxid=_mock_knoxid(request.user))
    completion = get_object_or_404(completions, id=completion_id)
    form = CompletionCodeForm(request.POST)
    if form.is_valid():
        code = form.cleaned_data["completion_code"]
        now = timezone.now()
        with transaction.atomic():
            completion.completion_code = code
            completion.is_completed = True
            completion.completed_at = now
            completion.save(
                update_fields=["completion_code", "is_completed", "completed_at", "updated_at"]
            )
            CompletionSubmissionLog.objects.create(
                completion=completion,
                submitted_by=request.user,
                completion_code=code,
            )
        messages.success(request, "수료코드가 등록되어 수료 처리되었습니다.")
    else:
        messages.error(request, "수료코드를 입력해 주세요.")
    return redirect(
        f"{reverse('training_dashboard')}?year={completion.target_year}"
    )


def _excel_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@login_required
@permission_required("chemicals.change_trainingcompletion", raise_exception=True)
def training_upload(request):
    form = TrainingUploadForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        try:
            workbook = load_workbook(
                form.cleaned_data["excel_file"], read_only=True, data_only=True
            )
            sheet = workbook.active
            header_row = next(sheet.iter_rows(values_only=True), ())
            headers = {_excel_text(value): index for index, value in enumerate(header_row)}
            required_headers = ("knoxid", "이름", "부서")
            missing_headers = [header for header in required_headers if header not in headers]
            if missing_headers:
                form.add_error(
                    "excel_file",
                    f"필수 열이 없습니다: {', '.join(missing_headers)}",
                )
            else:
                rows = []
                seen_knoxids = set()
                errors = []
                for row_number, row in enumerate(
                    sheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    values = {
                        header: _excel_text(row[headers[header]])
                        if headers[header] < len(row)
                        else ""
                        for header in required_headers
                    }
                    if not any(values.values()):
                        continue
                    empty_fields = [header for header, value in values.items() if not value]
                    if empty_fields:
                        errors.append(f"{row_number}행: {', '.join(empty_fields)} 값이 비어 있습니다.")
                    elif values["knoxid"] in seen_knoxids:
                        errors.append(f"{row_number}행: knoxid가 파일 안에서 중복됩니다.")
                    else:
                        seen_knoxids.add(values["knoxid"])
                        rows.append(values)

                if errors:
                    shown_errors = errors[:10]
                    if len(errors) > 10:
                        shown_errors.append(f"그 외 {len(errors) - 10}건")
                    form.add_error("excel_file", " ".join(shown_errors))
                elif not rows:
                    form.add_error("excel_file", "등록할 대상자 데이터가 없습니다.")
                else:
                    with transaction.atomic():
                        for row in rows:
                            handler, _ = HandlerProfile.objects.update_or_create(
                                knoxid=row["knoxid"],
                                defaults={
                                    "name": row["이름"],
                                    "department": row["부서"],
                                    "is_active": True,
                                },
                            )
                            TrainingCompletion.objects.get_or_create(
                                handler=handler,
                                target_year=form.cleaned_data["target_year"],
                            )
                    messages.success(
                        request,
                        f"{form.cleaned_data['target_year']}년 대상자 {len(rows)}명을 갱신했습니다.",
                    )
                    return redirect(
                        f"{reverse('training_dashboard')}?year={form.cleaned_data['target_year']}"
                    )
        except (BadZipFile, InvalidFileException, OSError, ValueError, StopIteration):
            form.add_error("excel_file", "엑셀 파일을 읽을 수 없습니다.")
        finally:
            if "workbook" in locals():
                workbook.close()

    return render(request, "chemicals/training_upload.html", {"form": form})


@login_required
@permission_required("chemicals.change_trainingcompletion", raise_exception=True)
def training_export_csv(request):
    target_year = _year_from_request(request)
    completions = _filtered_completions(request, target_year)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response.write("\ufeff")
    response["Content-Disposition"] = (
        f'attachment; filename="handler_training_{target_year}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(["대상 연도", "knoxid", "이름", "부서", "상태", "수료코드", "수료 처리 시각"])
    for completion in completions:
        writer.writerow(
            [
                completion.target_year,
                completion.handler.knoxid,
                completion.handler.name,
                completion.handler.department,
                "수료" if completion.is_completed else "미수료",
                completion.completion_code,
                timezone.localtime(completion.completed_at).strftime("%Y-%m-%d %H:%M")
                if completion.completed_at
                else "",
            ]
        )
    return response
